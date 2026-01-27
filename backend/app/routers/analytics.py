from datetime import datetime, timedelta
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import select, func, and_, or_, desc, extract, case

from ..db import get_db
from ..models import (
    User, Therapy, TherapySession, SessionStatus,
    CreditLedger, ActivityLog, DailyStats, Plan, UserPlan
)
from ..auth import require_superadmin

router = APIRouter(prefix="/api/analytics", tags=["analytics"])


def _iso(dt: datetime | None) -> str | None:
    return dt.isoformat() if dt else None


def _safe_div(n: float, d: float) -> float:
    return round(n / d, 2) if d else 0.0


# ──────────────────────────────────────────────────────────────
# Schemas
# ──────────────────────────────────────────────────────────────
class DashboardStats(BaseModel):
    total_users: int
    active_users_today: int
    active_users_week: int
    total_therapies: int
    total_sessions: int
    sessions_today: int
    sessions_week: int
    sessions_month: int
    credits_consumed_today: int
    credits_consumed_week: int
    credits_consumed_month: int
    credits_added_month: int
    avg_session_duration_min: float
    completion_rate: float


class TherapyUsageItem(BaseModel):
    therapy_id: int
    therapy_name: str
    category: str | None
    total_sessions: int
    completed_sessions: int
    total_duration_min: int
    credits_consumed: int
    avg_duration_min: float
    completion_rate: float


class UserActivityItem(BaseModel):
    user_id: int
    user_email: str
    user_name: str | None
    total_sessions: int
    total_duration_min: int
    credits_consumed: int
    credits_balance: int
    last_session: datetime | None


class TimeSeriesPoint(BaseModel):
    date: str
    value: int


class CreditFlowItem(BaseModel):
    date: str
    credits_added: int
    credits_consumed: int
    net_change: int


class SessionLogItem(BaseModel):
    id: int
    user_email: str
    user_name: str | None
    therapy_name: str
    started_at: datetime
    ended_at: datetime | None
    duration_min: int
    status: str
    credits_consumed: int


class ActivityLogItem(BaseModel):
    id: int
    user_email: str | None
    action: str
    description: str | None
    created_at: datetime
    ip_address: str | None


# ──────────────────────────────────────────────────────────────
# Helper Functions
# ──────────────────────────────────────────────────────────────
def get_date_range(period: str) -> tuple[datetime, datetime]:
    """Obtener rango de fechas según periodo."""
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    if period == "today":
        return today_start, now
    elif period == "week":
        week_start = today_start - timedelta(days=today_start.weekday())
        return week_start, now
    elif period == "month":
        month_start = today_start.replace(day=1)
        return month_start, now
    elif period == "quarter":
        quarter_month = ((today_start.month - 1) // 3) * 3 + 1
        quarter_start = today_start.replace(month=quarter_month, day=1)
        return quarter_start, now
    elif period == "year":
        year_start = today_start.replace(month=1, day=1)
        return year_start, now
    else:  # all time
        return datetime(2020, 1, 1), now


# ──────────────────────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────────────────────
@router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Estadísticas principales del dashboard."""
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)
    
    # Usuarios
    total_users = db.execute(select(func.count(User.id))).scalar() or 0
    
    # Usuarios activos (con sesiones)
    active_today = db.execute(
        select(func.count(func.distinct(TherapySession.user_id)))
        .where(TherapySession.started_at >= today_start)
    ).scalar() or 0
    
    active_week = db.execute(
        select(func.count(func.distinct(TherapySession.user_id)))
        .where(TherapySession.started_at >= week_start)
    ).scalar() or 0
    
    # Terapias
    total_therapies = db.execute(
        select(func.count(Therapy.id)).where(Therapy.is_active == True)
    ).scalar() or 0
    
    # Sesiones
    total_sessions = db.execute(select(func.count(TherapySession.id))).scalar() or 0
    
    sessions_today = db.execute(
        select(func.count(TherapySession.id))
        .where(TherapySession.started_at >= today_start)
    ).scalar() or 0
    
    sessions_week = db.execute(
        select(func.count(TherapySession.id))
        .where(TherapySession.started_at >= week_start)
    ).scalar() or 0
    
    sessions_month = db.execute(
        select(func.count(TherapySession.id))
        .where(TherapySession.started_at >= month_start)
    ).scalar() or 0
    
    # Créditos consumidos
    credits_today = db.execute(
        select(func.coalesce(func.sum(TherapySession.credits_consumed), 0))
        .where(TherapySession.started_at >= today_start)
    ).scalar() or 0
    
    credits_week = db.execute(
        select(func.coalesce(func.sum(TherapySession.credits_consumed), 0))
        .where(TherapySession.started_at >= week_start)
    ).scalar() or 0
    
    credits_month = db.execute(
        select(func.coalesce(func.sum(TherapySession.credits_consumed), 0))
        .where(TherapySession.started_at >= month_start)
    ).scalar() or 0
    
    # Créditos agregados este mes
    credits_added = db.execute(
        select(func.coalesce(func.sum(CreditLedger.delta), 0))
        .where(and_(
            CreditLedger.created_at >= month_start,
            CreditLedger.delta > 0
        ))
    ).scalar() or 0
    
    # Duración promedio
    avg_duration = db.execute(
        select(func.avg(TherapySession.duration_actual_sec))
        .where(TherapySession.status == SessionStatus.completed)
    ).scalar() or 0
    avg_duration_min = round(avg_duration / 60, 1) if avg_duration else 0
    
    # Tasa de completado
    completed = db.execute(
        select(func.count(TherapySession.id))
        .where(TherapySession.status == SessionStatus.completed)
    ).scalar() or 0
    completion_rate = round((completed / total_sessions * 100), 1) if total_sessions > 0 else 0
    
    return DashboardStats(
        total_users=total_users,
        active_users_today=active_today,
        active_users_week=active_week,
        total_therapies=total_therapies,
        total_sessions=total_sessions,
        sessions_today=sessions_today,
        sessions_week=sessions_week,
        sessions_month=sessions_month,
        credits_consumed_today=credits_today,
        credits_consumed_week=credits_week,
        credits_consumed_month=credits_month,
        credits_added_month=credits_added,
        avg_session_duration_min=avg_duration_min,
        completion_rate=completion_rate,
    )


@router.get("/therapies/usage", response_model=list[TherapyUsageItem])
async def get_therapy_usage(
    period: str = Query("month", regex="^(today|week|month|quarter|year|all)$"),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Uso de terapias por periodo."""
    start_date, end_date = get_date_range(period)
    
    # Subquery para estadísticas de sesiones
    stmt = (
        select(
            TherapySession.therapy_id,
            func.count(TherapySession.id).label("total_sessions"),
            func.sum(
                case((TherapySession.status == SessionStatus.completed, 1), else_=0)
            ).label("completed_sessions"),
            func.coalesce(func.sum(TherapySession.duration_actual_sec), 0).label("total_duration"),
            func.coalesce(func.sum(TherapySession.credits_consumed), 0).label("credits_consumed"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(TherapySession.therapy_id)
        .order_by(desc("total_sessions"))
        .limit(limit)
    )
    
    results = db.execute(stmt).all()
    
    items = []
    for row in results:
        therapy = db.get(Therapy, row.therapy_id)
        if not therapy:
            continue
        
        total = row.total_sessions or 0
        completed = row.completed_sessions or 0
        duration = row.total_duration or 0
        
        items.append(TherapyUsageItem(
            therapy_id=therapy.id,
            therapy_name=therapy.name,
            category=therapy.category,
            total_sessions=total,
            completed_sessions=completed,
            total_duration_min=duration // 60,
            credits_consumed=row.credits_consumed or 0,
            avg_duration_min=round(duration / total / 60, 1) if total > 0 else 0,
            completion_rate=round(completed / total * 100, 1) if total > 0 else 0,
        ))
    
    return items


@router.get("/users/activity", response_model=list[UserActivityItem])
async def get_user_activity(
    period: str = Query("month", regex="^(today|week|month|quarter|year|all)$"),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Actividad de usuarios por periodo."""
    start_date, end_date = get_date_range(period)
    
    stmt = (
        select(
            TherapySession.user_id,
            func.count(TherapySession.id).label("total_sessions"),
            func.coalesce(func.sum(TherapySession.duration_actual_sec), 0).label("total_duration"),
            func.coalesce(func.sum(TherapySession.credits_consumed), 0).label("credits_consumed"),
            func.max(TherapySession.started_at).label("last_session"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(TherapySession.user_id)
        .order_by(desc("total_sessions"))
        .limit(limit)
    )
    
    results = db.execute(stmt).all()
    
    items = []
    for row in results:
        user = db.get(User, row.user_id)
        if not user:
            continue
        
        items.append(UserActivityItem(
            user_id=user.id,
            user_email=user.email,
            user_name=user.name,
            total_sessions=row.total_sessions or 0,
            total_duration_min=(row.total_duration or 0) // 60,
            credits_consumed=row.credits_consumed or 0,
            credits_balance=user.credits_balance,
            last_session=row.last_session,
        ))
    
    return items


@router.get("/sessions/timeline", response_model=list[TimeSeriesPoint])
async def get_sessions_timeline(
    period: str = Query("month", regex="^(week|month|quarter|year)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Serie temporal de sesiones por día."""
    start_date, end_date = get_date_range(period)
    
    stmt = (
        select(
            func.date(TherapySession.started_at).label("date"),
            func.count(TherapySession.id).label("count"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(func.date(TherapySession.started_at))
        .order_by("date")
    )
    
    results = db.execute(stmt).all()
    
    return [
        TimeSeriesPoint(date=str(row.date), value=row.count)
        for row in results
    ]


@router.get("/credits/flow", response_model=list[CreditFlowItem])
async def get_credits_flow(
    period: str = Query("month", regex="^(week|month|quarter|year)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Flujo de créditos (agregados vs consumidos) por día."""
    start_date, end_date = get_date_range(period)
    
    # Créditos agregados por día
    added_stmt = (
        select(
            func.date(CreditLedger.created_at).label("date"),
            func.coalesce(func.sum(
                case((CreditLedger.delta > 0, CreditLedger.delta), else_=0)
            ), 0).label("added"),
            func.coalesce(func.sum(
                case((CreditLedger.delta < 0, func.abs(CreditLedger.delta)), else_=0)
            ), 0).label("consumed"),
        )
        .where(CreditLedger.created_at.between(start_date, end_date))
        .group_by(func.date(CreditLedger.created_at))
        .order_by("date")
    )
    
    results = db.execute(added_stmt).all()
    
    return [
        CreditFlowItem(
            date=str(row.date),
            credits_added=row.added,
            credits_consumed=row.consumed,
            net_change=row.added - row.consumed,
        )
        for row in results
    ]


@router.get("/sessions/recent", response_model=list[SessionLogItem])
async def get_recent_sessions(
    limit: int = Query(50, ge=1, le=200),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Sesiones recientes con detalles."""
    stmt = (
        select(TherapySession)
        .order_by(desc(TherapySession.started_at))
        .limit(limit)
    )
    
    if status:
        stmt = stmt.where(TherapySession.status == status)
    
    sessions = db.execute(stmt).scalars().all()
    
    items = []
    for s in sessions:
        user = db.get(User, s.user_id)
        therapy = db.get(Therapy, s.therapy_id)
        
        items.append(SessionLogItem(
            id=s.id,
            user_email=user.email if user else "Unknown",
            user_name=user.name if user else None,
            therapy_name=therapy.name if therapy else "Unknown",
            started_at=s.started_at,
            ended_at=s.ended_at,
            duration_min=s.duration_actual_sec // 60,
            status=s.status.value,
            credits_consumed=s.credits_consumed,
        ))
    
    return items


@router.get("/activity/log", response_model=list[ActivityLogItem])
async def get_activity_log(
    limit: int = Query(100, ge=1, le=500),
    action: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Log de actividad del sistema."""
    stmt = (
        select(ActivityLog)
        .order_by(desc(ActivityLog.created_at))
        .limit(limit)
    )
    
    if action:
        stmt = stmt.where(ActivityLog.action == action)
    
    logs = db.execute(stmt).scalars().all()
    
    items = []
    for log in logs:
        user = db.get(User, log.user_id) if log.user_id else None
        
        items.append(ActivityLogItem(
            id=log.id,
            user_email=user.email if user else None,
            action=log.action,
            description=log.description,
            created_at=log.created_at,
            ip_address=log.ip_address,
        ))
    
    return items


@router.get("/categories/distribution")
async def get_category_distribution(
    period: str = Query("month", regex="^(today|week|month|quarter|year|all)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Distribución de sesiones por categoría de terapia."""
    start_date, end_date = get_date_range(period)
    
    stmt = (
        select(
            Therapy.category,
            func.count(TherapySession.id).label("count"),
        )
        .join(Therapy, TherapySession.therapy_id == Therapy.id)
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(Therapy.category)
        .order_by(desc("count"))
    )
    
    results = db.execute(stmt).all()
    
    return [
        {"category": row.category or "Sin categoría", "count": row.count}
        for row in results
    ]


@router.get("/hours/distribution")
async def get_hours_distribution(
    period: str = Query("month", regex="^(today|week|month|quarter|year|all)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Distribución de sesiones por hora del día."""
    start_date, end_date = get_date_range(period)
    
    stmt = (
        select(
            extract("hour", TherapySession.started_at).label("hour"),
            func.count(TherapySession.id).label("count"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by("hour")
        .order_by("hour")
    )
    
    results = db.execute(stmt).all()
    
    # Llenar horas faltantes con 0
    hour_counts = {int(row.hour): row.count for row in results}
    return [
        {"hour": h, "count": hour_counts.get(h, 0)}
        for h in range(24)
    ]


@router.get("/export/sessions")
async def export_sessions(
    start_date: datetime = Query(...),
    end_date: datetime = Query(...),
    format: str = Query("json", regex="^(json|csv)$"),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Exportar sesiones para análisis externo."""
    stmt = (
        select(TherapySession)
        .where(TherapySession.started_at.between(start_date, end_date))
        .order_by(TherapySession.started_at)
    )
    
    sessions = db.execute(stmt).scalars().all()
    
    data = []
    for s in sessions:
        user = db.get(User, s.user_id)
        therapy = db.get(Therapy, s.therapy_id)
        
        data.append({
            "id": s.id,
            "user_email": user.email if user else None,
            "user_name": user.name if user else None,
            "therapy_name": therapy.name if therapy else None,
            "therapy_category": therapy.category if therapy else None,
            "started_at": s.started_at.isoformat(),
            "ended_at": s.ended_at.isoformat() if s.ended_at else None,
            "duration_planned_sec": s.duration_planned_sec,
            "duration_actual_sec": s.duration_actual_sec,
            "status": s.status.value,
            "credits_consumed": s.credits_consumed,
            "arduino_connected": s.arduino_connected,
            "color_mode_used": s.color_mode_used,
        })
    
    if format == "csv":
        import csv
        import io
        from fastapi.responses import StreamingResponse
        
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=sessions_{start_date.date()}_{end_date.date()}.csv"}
        )
    
    return {"sessions": data, "count": len(data)}


@router.get("/export/report")
async def export_analytics_report(
    start_date: datetime = Query(...),
    end_date: datetime = Query(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_superadmin),
):
    """Exportación profesional de analíticas (Excel con múltiples hojas y gráficas)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.chart import LineChart, BarChart, Reference
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except Exception:
        raise HTTPException(status_code=500, detail="Dependencia openpyxl no disponible")

    # ── KPIs principales
    total_users = db.execute(select(func.count(User.id))).scalar() or 0
    total_therapies = db.execute(select(func.count(Therapy.id))).scalar() or 0
    total_sessions = db.execute(
        select(func.count(TherapySession.id))
        .where(TherapySession.started_at.between(start_date, end_date))
    ).scalar() or 0
    active_users = db.execute(
        select(func.count(func.distinct(TherapySession.user_id)))
        .where(TherapySession.started_at.between(start_date, end_date))
    ).scalar() or 0
    completed_sessions = db.execute(
        select(func.count(TherapySession.id))
        .where(
            and_(
                TherapySession.started_at.between(start_date, end_date),
                TherapySession.status == SessionStatus.completed,
            )
        )
    ).scalar() or 0
    credits_added = db.execute(
        select(func.coalesce(func.sum(CreditLedger.delta), 0))
        .where(
            and_(
                CreditLedger.created_at.between(start_date, end_date),
                CreditLedger.delta > 0,
            )
        )
    ).scalar() or 0
    credits_consumed = db.execute(
        select(func.coalesce(func.sum(TherapySession.credits_consumed), 0))
        .where(TherapySession.started_at.between(start_date, end_date))
    ).scalar() or 0
    avg_duration = db.execute(
        select(func.avg(TherapySession.duration_actual_sec))
        .where(
            and_(
                TherapySession.started_at.between(start_date, end_date),
                TherapySession.status == SessionStatus.completed,
            )
        )
    ).scalar() or 0
    avg_duration_min = round((avg_duration or 0) / 60, 2)

    # ── Time series sesiones
    timeline_stmt = (
        select(
            func.date(TherapySession.started_at).label("date"),
            func.count(TherapySession.id).label("count"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(func.date(TherapySession.started_at))
        .order_by("date")
    )
    timeline = db.execute(timeline_stmt).all()

    # ── Flow créditos
    credits_stmt = (
        select(
            func.date(CreditLedger.created_at).label("date"),
            func.coalesce(func.sum(case((CreditLedger.delta > 0, CreditLedger.delta), else_=0)), 0).label("added"),
            func.coalesce(func.sum(case((CreditLedger.delta < 0, func.abs(CreditLedger.delta)), else_=0)), 0).label("consumed"),
        )
        .where(CreditLedger.created_at.between(start_date, end_date))
        .group_by(func.date(CreditLedger.created_at))
        .order_by("date")
    )
    credits_flow = db.execute(credits_stmt).all()

    # ── Top terapias
    usage_stmt = (
        select(
            TherapySession.therapy_id,
            func.count(TherapySession.id).label("total_sessions"),
            func.coalesce(func.sum(TherapySession.duration_actual_sec), 0).label("total_duration"),
            func.coalesce(func.sum(TherapySession.credits_consumed), 0).label("credits_consumed"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(TherapySession.therapy_id)
        .order_by(desc("total_sessions"))
        .limit(20)
    )
    usage_rows = db.execute(usage_stmt).all()

    # ── Top usuarios
    users_stmt = (
        select(
            TherapySession.user_id,
            func.count(TherapySession.id).label("total_sessions"),
            func.coalesce(func.sum(TherapySession.duration_actual_sec), 0).label("total_duration"),
            func.coalesce(func.sum(TherapySession.credits_consumed), 0).label("credits_consumed"),
            func.max(TherapySession.started_at).label("last_session"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(TherapySession.user_id)
        .order_by(desc("total_sessions"))
        .limit(20)
    )
    top_users = db.execute(users_stmt).all()

    # ── Distribución por categoría
    categories_stmt = (
        select(
            Therapy.category,
            func.count(TherapySession.id).label("count"),
        )
        .join(Therapy, TherapySession.therapy_id == Therapy.id)
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by(Therapy.category)
        .order_by(desc("count"))
    )
    categories_dist = db.execute(categories_stmt).all()

    # ── Distribución por hora
    hours_stmt = (
        select(
            extract("hour", TherapySession.started_at).label("hour"),
            func.count(TherapySession.id).label("count"),
        )
        .where(TherapySession.started_at.between(start_date, end_date))
        .group_by("hour")
        .order_by("hour")
    )
    hours_dist = db.execute(hours_stmt).all()

    # ── Sesiones detalle
    sessions = db.execute(
        select(TherapySession)
        .where(TherapySession.started_at.between(start_date, end_date))
        .order_by(TherapySession.started_at)
    ).scalars().all()

    # ── Workbook
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"

    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111827")
    accent_fill = PatternFill("solid", fgColor="1F2937")

    ws_summary["A1"] = "Reporte Profesional de Analíticas"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Periodo: {start_date.date()} → {end_date.date()}"
    ws_summary["A3"] = f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

    summary_rows = [
        ("Usuarios totales", total_users),
        ("Usuarios activos (periodo)", active_users),
        ("Terapias totales", total_therapies),
        ("Sesiones totales (periodo)", total_sessions),
        ("Sesiones completadas", completed_sessions),
        ("Tasa de completado %", _safe_div(completed_sessions * 100, total_sessions)),
        ("Créditos agregados", credits_added),
        ("Créditos consumidos", credits_consumed),
        ("Balance neto", credits_added - credits_consumed),
        ("Duración promedio (min)", avg_duration_min),
    ]

    ws_summary.append(["Métrica", "Valor"])
    ws_summary["A5"].font = header_font
    ws_summary["B5"].font = header_font
    ws_summary["A5"].fill = header_fill
    ws_summary["B5"].fill = header_fill
    ws_summary["A5"].alignment = Alignment(horizontal="center")
    ws_summary["B5"].alignment = Alignment(horizontal="center")

    for label, value in summary_rows:
        ws_summary.append([label, value])

    ws_summary.freeze_panes = "A6"

    for col in range(1, 3):
        ws_summary.column_dimensions[get_column_letter(col)].width = 32

    # ── Hoja: Sesiones timeline
    ws_timeline = wb.create_sheet("Sesiones (timeline)")
    ws_timeline.append(["Fecha", "Sesiones"])
    ws_timeline["A1"].font = header_font
    ws_timeline["B1"].font = header_font
    ws_timeline["A1"].fill = header_fill
    ws_timeline["B1"].fill = header_fill
    ws_timeline.freeze_panes = "A2"

    for row in timeline:
        ws_timeline.append([str(row.date), int(row.count)])

    chart = LineChart()
    chart.title = "Sesiones por día"
    chart.y_axis.title = "Sesiones"
    chart.x_axis.title = "Fecha"
    data = Reference(ws_timeline, min_col=2, min_row=1, max_row=ws_timeline.max_row)
    cats = Reference(ws_timeline, min_col=1, min_row=2, max_row=ws_timeline.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_timeline.add_chart(chart, "D2")

    # ── Hoja: Créditos
    ws_credits = wb.create_sheet("Créditos")
    ws_credits.append(["Fecha", "Créditos agregados", "Créditos consumidos", "Net"])
    for col in range(1, 5):
        cell = ws_credits.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    ws_credits.freeze_panes = "A2"

    for row in credits_flow:
        added = int(row.added or 0)
        consumed = int(row.consumed or 0)
        ws_credits.append([str(row.date), added, consumed, added - consumed])

    credits_chart = LineChart()
    credits_chart.title = "Flujo de créditos"
    credits_chart.y_axis.title = "Créditos"
    credits_chart.x_axis.title = "Fecha"
    data = Reference(ws_credits, min_col=2, min_row=1, max_col=4, max_row=ws_credits.max_row)
    cats = Reference(ws_credits, min_col=1, min_row=2, max_row=ws_credits.max_row)
    credits_chart.add_data(data, titles_from_data=True)
    credits_chart.set_categories(cats)
    ws_credits.add_chart(credits_chart, "F2")

    # ── Hoja: Top Terapias
    ws_therapies = wb.create_sheet("Top terapias")
    ws_therapies.append(["Terapia", "Categoría", "Sesiones", "Duración total (min)", "Créditos"])
    for col in range(1, 6):
        cell = ws_therapies.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    ws_therapies.freeze_panes = "A2"

    for row in usage_rows:
        therapy = db.get(Therapy, row.therapy_id)
        if not therapy:
            continue
        ws_therapies.append([
            therapy.name,
            therapy.category or "Sin categoría",
            int(row.total_sessions or 0),
            int((row.total_duration or 0) // 60),
            int(row.credits_consumed or 0),
        ])

    bar = BarChart()
    bar.title = "Top terapias por sesiones"
    bar.y_axis.title = "Sesiones"
    bar.x_axis.title = "Terapia"
    data = Reference(ws_therapies, min_col=3, min_row=1, max_row=ws_therapies.max_row)
    cats = Reference(ws_therapies, min_col=1, min_row=2, max_row=ws_therapies.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_therapies.add_chart(bar, "G2")

    # ── Hoja: Top usuarios
    ws_users = wb.create_sheet("Top usuarios")
    ws_users.append(["Usuario", "Email", "Sesiones", "Duración total (min)", "Créditos", "Última sesión"])
    for col in range(1, 7):
        cell = ws_users.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    ws_users.freeze_panes = "A2"

    for row in top_users:
        user = db.get(User, row.user_id)
        ws_users.append([
            user.name if user else None,
            user.email if user else None,
            int(row.total_sessions or 0),
            int((row.total_duration or 0) // 60),
            int(row.credits_consumed or 0),
            _iso(row.last_session),
        ])

    users_bar = BarChart()
    users_bar.title = "Top usuarios por sesiones"
    users_bar.y_axis.title = "Sesiones"
    users_bar.x_axis.title = "Usuario"
    data = Reference(ws_users, min_col=3, min_row=1, max_row=ws_users.max_row)
    cats = Reference(ws_users, min_col=1, min_row=2, max_row=ws_users.max_row)
    users_bar.add_data(data, titles_from_data=True)
    users_bar.set_categories(cats)
    ws_users.add_chart(users_bar, "H2")

    # ── Hoja: Categorías
    ws_categories = wb.create_sheet("Categorías")
    ws_categories.append(["Categoría", "Sesiones"])
    for col in range(1, 3):
        cell = ws_categories.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    ws_categories.freeze_panes = "A2"

    for row in categories_dist:
        ws_categories.append([row.category or "Sin categoría", int(row.count or 0)])

    cat_bar = BarChart()
    cat_bar.title = "Sesiones por categoría"
    cat_bar.y_axis.title = "Sesiones"
    cat_bar.x_axis.title = "Categoría"
    data = Reference(ws_categories, min_col=2, min_row=1, max_row=ws_categories.max_row)
    cats = Reference(ws_categories, min_col=1, min_row=2, max_row=ws_categories.max_row)
    cat_bar.add_data(data, titles_from_data=True)
    cat_bar.set_categories(cats)
    ws_categories.add_chart(cat_bar, "D2")

    # ── Hoja: Horas
    ws_hours = wb.create_sheet("Horas")
    ws_hours.append(["Hora", "Sesiones"])
    for col in range(1, 3):
        cell = ws_hours.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    ws_hours.freeze_panes = "A2"

    hour_counts = {int(r.hour): int(r.count or 0) for r in hours_dist}
    for h in range(24):
        ws_hours.append([h, hour_counts.get(h, 0)])

    hours_line = LineChart()
    hours_line.title = "Sesiones por hora"
    hours_line.y_axis.title = "Sesiones"
    hours_line.x_axis.title = "Hora"
    data = Reference(ws_hours, min_col=2, min_row=1, max_row=ws_hours.max_row)
    cats = Reference(ws_hours, min_col=1, min_row=2, max_row=ws_hours.max_row)
    hours_line.add_data(data, titles_from_data=True)
    hours_line.set_categories(cats)
    ws_hours.add_chart(hours_line, "D2")

    # ── Hoja: Sesiones (detalle)
    ws_sessions = wb.create_sheet("Sesiones (detalle)")
    ws_sessions.append([
        "ID", "Usuario", "Email", "Terapia", "Categoría", "Inicio", "Fin", "Duración plan (seg)",
        "Duración real (seg)", "Estado", "Créditos", "Arduino", "Color"
    ])
    for col in range(1, 14):
        cell = ws_sessions.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    ws_sessions.freeze_panes = "A2"

    for s in sessions:
        user = db.get(User, s.user_id)
        therapy = db.get(Therapy, s.therapy_id)
        ws_sessions.append([
            s.id,
            user.name if user else None,
            user.email if user else None,
            therapy.name if therapy else None,
            therapy.category if therapy else None,
            _iso(s.started_at),
            _iso(s.ended_at),
            s.duration_planned_sec,
            s.duration_actual_sec,
            s.status.value,
            s.credits_consumed,
            s.arduino_connected,
            s.color_mode_used,
        ])

    # Ajustar ancho de columnas
    for ws in [ws_timeline, ws_credits, ws_therapies, ws_users, ws_categories, ws_hours, ws_sessions]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    # Formato numérico y filtros
    ws_sessions.auto_filter.ref = ws_sessions.dimensions
    ws_therapies.auto_filter.ref = ws_therapies.dimensions
    ws_users.auto_filter.ref = ws_users.dimensions
    ws_categories.auto_filter.ref = ws_categories.dimensions

    # Resaltar valores altos en hojas clave
    ws_therapies.conditional_formatting.add(
        f"C2:C{ws_therapies.max_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["10"], fill=accent_fill),
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"analytics_report_{start_date.date()}_{end_date.date()}.xlsx"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
